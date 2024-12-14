import easyocr
import openpyxl

# Tạo đối tượng Reader của EasyOCR và chỉ định sử dụng GPU nếu bạn có
reader = easyocr.Reader(['vi'], gpu=True)

# Đọc văn bản từ hình ảnh
image_path = 'dulieu.jpg'
result = reader.readtext(image_path)

# Tạo workbook và worksheet mới
wb = openpyxl.Workbook()
ws = wb.active

# Chuyển kết quả đọc được thành danh sách và ghi vào worksheet
data = []
for (_, text, _) in result:
    # Giả sử mỗi dòng trong bảng được tách bằng khoảng trắng
    data.append(text.split())

# Kết quả đầu ra với các phần tử đã được xử lý
processed_data = []

for item in data:
    if isinstance(item, list):
        # Nối các chữ trong list lại thành một đoạn text và thay dấu gạch ngang bằng khoảng cách
        processed_item = " ".join(item).replace("_", " ")
        processed_data.append(processed_item)
    else:
        # Giữ nguyên nếu không phải là list
        processed_data.append(item)

# print("tiêu đề", processed_data[0])
# print("data", processed_data[1:])
data = processed_data[1:]
# print(data)
# Số lượng phần tử trong mỗi danh sách con
n = 7
# Ghép mỗi 7 phần tử thành một list con
list_2d = [data[i:i+n] for i in range(0, len(data), n)]

# Chuyển các giá trị số thành số nguyên
for row in list_2d:
    for i in range(len(row)):
        try:
            row[i] = int(row[i])
        except ValueError:
            pass  # Nếu không thể chuyển đổi thì bỏ qua

print(list_2d)

# Thêm chuỗi "tiêu đề" vào đầu danh sách
list_2d.insert(0, [processed_data[0]])

# Ghi dữ liệu vào worksheet
for row in list_2d:
    ws.append(row)

# Nối 7 ô đầu tiên thành 1 ô để chứa chữ "đầu dòng"
ws.merge_cells('A1:G1')
ws['A1'].value = processed_data[0]

# Định dạng ô hợp nhất để căn giữa
ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

# Lưu file Excel
wb.save('danhsach.xlsx')
